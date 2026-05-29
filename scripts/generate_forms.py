"""
Generiert m/f/n-Triplets für Adjektive und Genus/Plural/Gen.Sg für Nomen.

Strategie:
- Heuristik für regelmäßige Fälle (Endungs-basiert).
- Override-Maps für Sonderfälle: bewegliche Vokale, Stammwechsel,
  feminine i-Deklination, maskuline Personennamen auf -a.
- Output nach data/forms.json (merged mit existierenden Daten).

Bekannte Limitationen:
- Heuristik trifft ca. 90-95% richtig. Restliche Fehler via Fehler-Knopf
  in der App melden und in der Override-Liste hier pflegen.
- Genitiv-Singular nur bei kuratierten beweglichen-Vokal-Fällen,
  weil heuristisch nicht zuverlässig erkennbar.
- Plural-Hinweis nur bei wirklich unregelmäßigen Stämmen, regulär
  Plurale werden weggelassen um Karten-Clutter zu vermeiden.
"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook


# =====================================================================
# ADJEKTIV-OVERRIDES
# =====================================================================
ADJ_OVERRIDE = {
    "roze": ["roze", "roze", "roze"],
    "topao": ["topao", "topla", "toplo"],
    "okrugao": ["okrugao", "okrugla", "okruglo"],
    "svetao": ["svetao", "svetla", "svetlo"],
    "kiseo": ["kiseo", "kisela", "kiselo"],
    "zreo": ["zreo", "zrela", "zrelo"],
    "beo": ["beo", "bela", "belo"],
    "radoznao": ["radoznao", "radoznala", "radoznalo"],
    "preostao": ["preostao", "preostala", "preostalo"],
    "odrastao": ["odrastao", "odrasla", "odraslo"],
    "bolestan": ["bolestan", "bolesna", "bolesno"],
    "koristan": ["koristan", "korisna", "korisno"],
    "radostan": ["radostan", "radosna", "radosno"],
    "svestan": ["svestan", "svesna", "svesno"],
    "zao": ["zao", "zla", "zlo"],
    "ceo": ["ceo", "cela", "celo"],
    "veseo": ["veseo", "vesela", "veselo"],
    "debeo": ["debeo", "debela", "debelo"],
    "smeo": ["smeo", "smela", "smelo"],
    "sladak": ["sladak", "slatka", "slatko"],
    "gladak": ["gladak", "glatka", "glatko"],
    "nizak": ["nizak", "niska", "nisko"],
    "redak": ["redak", "retka", "retko"],
    "kratak": ["kratak", "kratka", "kratko"],
    "težak": ["težak", "teška", "teško"],
    "blizak": ["blizak", "bliska", "blisko"],
    "uzak": ["uzak", "uska", "usko"],
    "električni": ["električni", "električna", "električno"],
    "kućni": ["kućni", "kućna", "kućno"],
    "lični": ["lični", "lična", "lično"],
    "mali": ["mali", "mala", "malo"],
    "dupli": ["dupli", "dupla", "duplo"],
    "isti": ["isti", "ista", "isto"],
    "opšti": ["opšti", "opšta", "opšte"],
    "poslovni": ["poslovni", "poslovna", "poslovno"],
    "međunarodni": ["međunarodni", "međunarodna", "međunarodno"],
    "kulturni": ["kulturni", "kulturna", "kulturno"],
    "naučni": ["naučni", "naučna", "naučno"],
    "raniji": ["raniji", "ranija", "ranije"],
    "bolji": ["bolji", "bolja", "bolje"],
    "važeći": ["važeći", "važeća", "važeće"],
    "iznenađujući": ["iznenađujući", "iznenađujuća", "iznenađujuće"],
    "odgovarajući": ["odgovarajući", "odgovarajuća", "odgovarajuće"],
    "budući": ["budući", "buduća", "buduće"],
    "lep": ["lep", "lepa", "lepo"],
    "jak": ["jak", "jaka", "jako"],
    "suv": ["suv", "suva", "suvo"],
    "živ": ["živ", "živa", "živo"],
    "njihov": ["njihov", "njihova", "njihovo"],
}

VOWELS = set("aeiouAEIOU")


def adj_has_movable_a(word):
    if len(word) < 4:
        return False
    if word[-1] not in "knrlmcčćš":
        return False
    if word[-2] != "a":
        return False
    if word[-3] in VOWELS:
        return False
    return True


def gen_adj_mfn(word):
    if word in ADJ_OVERRIDE:
        return ADJ_OVERRIDE[word]
    if word.endswith("ao") and len(word) >= 4:
        stem = word[:-2] + "l"
        return [word, stem + "a", stem + "o"]
    if word.endswith("eo") and len(word) >= 3:
        stem = word[:-2] + "el"
        return [word, stem + "a", stem + "o"]
    if word.endswith("i") and len(word) >= 3:
        stem = word[:-1]
        if word.endswith("ji") or word.endswith("ći"):
            return [word, stem + "a", stem + "e"]
        return [word, stem + "a", stem + "o"]
    if word.endswith("e") and len(word) <= 5:
        return [word, word, word]
    if adj_has_movable_a(word):
        stem = word[:-2] + word[-1]
        return [word, stem + "a", stem + "o"]
    return [word, word + "a", word + "o"]


# =====================================================================
# NOMEN-OVERRIDES
# =====================================================================

# Maskulin endend auf -a (Verwandtschaft, Anrede, slawische Personenwörter)
MASC_ON_A = {
    "tata", "deda", "kolega", "sluga", "vojvoda", "starešina",
    "vladika", "gazda", "čika", "papa", "Sava",
}

# Maskulin endend auf -o (Lehnwörter und l-Stamm-Wechsel)
# l-Stamm: posao → posla, udeo → udela (im Maskulin l→o)
MASC_ON_O = {
    "auto": None,
    "metro": None,
    "posao": "posla",   # Gen.Sg
    "udeo": "udela",
}

# Pluralia tantum: nur Plural existent, Genus ist klassifizierend
NEUT_PL_TANTUM = {"vrata", "leđa", "kola"}
FEM_PL_TANTUM = {
    "naočare", "naočale",
    "pantalone", "farmerke",
    "makaze", "makazice",
    "merdevine", "stepenice",
    "novine", "studije",
    # Plural-Formen femininer -a Wörter (Robin's xlsx listet sie im Pl.)
    "rukavice", "papuče", "čarape", "patike", "cipele", "čizme",
}


# =====================================================================
# VERB-OVERRIDES (1. Person Singular Präsens)
# =====================================================================

# Override für Verben, deren 1.P.Sg. nicht aus dem Infinitiv ableitbar ist.
# Key: Verb im Infinitiv (ohne reflexives "se").
VERB_OVERRIDE = {
    # ---- Echte Unregelmäßige ----
    "biti": "sam",
    "hteti": "hoću",
    "moći": "mogu",   # Spezial mit -u Endung
    "ići": "idem",
    "jesti": "jedem",
    "piti": "pijem",
    "dati": "dam",
    "znati": "znam",
    "stati": "stanem",
    "imati": "imam",
    "spavati": "spavam",
    "uzeti": "uzmem",
    "umeti": "umem",

    # ---- -ći Verben (Velarwechsel und/oder Stammwechsel) ----
    "doći": "dođem",
    "naći": "nađem",
    "preći": "pređem",
    "ući": "uđem",
    "izaći": "izađem",
    "izići": "iziđem",
    "stići": "stignem",
    "leći": "legnem",
    "reći": "kažem",      # kompletter Stammwechsel
    "peći": "pečem",      # k→č
    "seći": "sečem",      # k→č
    "obući": "obučem",
    "izvući": "izvučem",
    "vući": "vučem",
    "tući": "tučem",

    # ---- Verben auf -sti (Stammerweiterung) ----
    "sresti": "sretnem",
    "rasti": "rastem",
    "pasti": "padnem",
    "sesti": "sednem",
    "naprasti": "naprestanem",

    # ---- -ati mit Konsonantenwechsel ----
    "pisati": "pišem",
    "kazati": "kažem",
    "brisati": "brišem",
    "plakati": "plačem",
    "skakati": "skačem",
    "iskakati": "iskačem",
    "tražiti": "tražim",   # regulär eigentlich
    "vezati": "vežem",
    "lagati": "lažem",

    # ---- -avati Klasse 1 (verkürzt, -ajem) ----
    "davati": "dajem",
    "prodavati": "prodajem",
    "predavati": "predajem",
    "saznavati": "saznajem",
    "izdavati": "izdajem",
    "ostavati": "ostajem",
    "udavati": "udajem",
    "postajati": "postajem",
    "ustajati": "ustajem",
    "nestajati": "nestajem",
    "smetati": "smetam",   # regulär -ati → -am

    # ---- -ati irreguläre Klassiker ----
    "zvati": "zovem",       # Stamm-o
    "brati": "berem",       # Stamm-e
    "prati": "perem",
    "spavati": "spavam",
    "smeti": "smem",
    "videti": "vidim",
    "želeti": "želim",
    "voleti": "volim",
    "umeti": "umem",
    "razumeti": "razumem",  # -em statt -im!
    "leteti": "letim",

    # ---- Perfektive Formen mit Stammerweiterung -ne- ----
    "postati": "postanem",
    "ostati": "ostanem",
    "stati": "stanem",
    "ustati": "ustanem",
    "nestati": "nestanem",
    "sresti": "sretnem",
    "propasti": "propadnem",
    "pasti": "padnem",
    "sesti": "sednem",

    # ---- Perfektive mit Stamm-Wechsel (Konsonant) ----
    "pokazati": "pokažem",
    "kazati": "kažem",
    "napisati": "napišem",
    "potpisati": "potpišem",
    "zaplakati": "zaplačem",
    "vezati": "vežem",
    "povezati": "povežem",
    "preskočiti": "preskočim",  # regulär

    # ---- Perfektive von -piti/-jesti/-nositi (Erweiterung) ----
    "popiti": "popijem",
    "pojesti": "pojedem",
    "doneti": "donesem",
    "odneti": "odnesem",
    "preneti": "prenesem",
    "uneti": "unesem",
    "izneti": "iznesem",
    "doneti se": "donesem se",

    # ---- Perfektive von zvati (Stamm zov-) ----
    "pozvati": "pozovem",
    "nazvati": "nazovem",
    "sazvati": "sazovem",

    # ---- Perfektive von -uzeti ----
    "zauzeti": "zauzmem",
    "preuzeti": "preuzmem",
    "oduzeti": "oduzmem",
    "preduzeti": "preduzmem",

    # ---- Perfektive von -znati ----
    "saznati": "saznam",
    "doznati": "doznam",
    "upoznati": "upoznam",

    # ---- Perfektive von -moći ----
    "pomoći": "pomognem",
    "omogućiti": "omogućim",

    # ---- Perfektive von -dati ----
    "predati": "predam",
    "izdati": "izdam",
    "udati se": "udam se",
    "prodati": "prodam",
    "dodati": "dodam",
    "zadati": "zadam",
    "podati se": "podam se",

    # ---- Diverse Klassiker ----
    "biti se": "bijem se",       # selten
    "objasniti": "objasnim",
    "razumeti": "razumem",
    "porazumeti se": "porazumem se",

    # ---- "Alte" -vati Verben (kein Stammwechsel zu -ujem) ----
    # Linguistisch: nicht Suffix-Erweiterung, sondern alter -vati Stamm
    "pozivati": "pozivam",
    "zalivati": "zalivam",
    "pokrivati": "pokrivam",
    "uživati": "uživam",
    "nazivati": "nazivam",
}


def gen_verb_pres1(word):
    """Generiert 1. P. Sg. Präsens. None wenn nicht ableitbar."""
    if not word:
        return None
    # Override hat Vorrang
    if word in VERB_OVERRIDE:
        return VERB_OVERRIDE[word]

    # Reflexive Verben: "se" ist NICHT Teil des Stamms, wir kriegen den Infinitiv ohne se
    # (wird vor Aufruf gestrippt)

    # -ovati → -ujem (regulär, Stammwechsel)
    if word.endswith("ovati") and len(word) > 5:
        return word[:-5] + "ujem"

    # -ivati → -ujem (regulär, Stammwechsel)
    if word.endswith("ivati") and len(word) > 5:
        return word[:-5] + "ujem"

    # -avati → -avam (default, neue Klasse). Klasse 1 (-ajem) ist in Override.
    if word.endswith("avati") and len(word) > 5:
        return word[:-5] + "avam"

    # -nuti → -nem
    if word.endswith("nuti") and len(word) > 4:
        return word[:-4] + "nem"

    # -iti → -im
    if word.endswith("iti") and len(word) > 3:
        return word[:-3] + "im"

    # -eti → -im (mit Ausnahmen in Override für -em wie razumeti)
    if word.endswith("eti") and len(word) > 3:
        return word[:-3] + "im"

    # -ati → -am (mit Ausnahmen für Konsonantenwechsel in Override)
    if word.endswith("ati") and len(word) > 3:
        return word[:-3] + "am"

    # Sonst: keine Heuristik
    return None


def strip_se(word):
    """Reflexives 'se' am Ende entfernen (gibt Tuple zurück mit reflexive-Flag)."""
    if word.endswith(" se"):
        return word[:-3].strip(), True
    return word, False

# Feminin endend auf Konsonant (i-Deklination, Abstrakta auf -ost/-nost und Klassiker)
FEM_ON_CONS = {
    "noć", "stvar", "smrt", "reč", "ljubav", "vlast", "čast", "krv",
    "peć", "sol", "jesen", "moć", "pomoć", "vest", "mast", "kost",
    "narav", "savest", "starost", "mladost", "smelost", "vrednost",
    "stvarnost", "stranica",
    # Klassiker abseits -ost
    "stvar", "nit", "stvar", "želja", "smrt",
}

# Neutrum-Ausnahmen auf Konsonant (selten: -e/-ena Stämme)
# In Serbisch eher selten, lass ich aus.

# Pluralform-Wechsel (nur unregelmäßige eintragen)
NOUN_PLURAL = {
    "čovek": "ljudi",
    "dete": "deca",
    "brat": "braća",
    "gospodin": "gospoda",
    "oko": "oči",
    "uho": "uši",
    "uvo": "uši",
    "vlastelin": "vlastela",
    "knez": "kneževi",
    "vlas": "vlasi",
    "Srbin": "Srbi",
    "građanin": "građani",
    "seljanin": "seljani",
    "ostrvljanin": "ostrvljani",
    "katolik": "katolici",
    "neprijatelj": "neprijatelji",
    "prijatelj": "prijatelji",
}

# Genitiv-Singular bei beweglichem Vokal (kuratiert)
NOUN_GEN = {
    "otac": "oca",
    "pas": "psa",
    "san": "sna",
    "vetar": "vetra",
    "novac": "novca",
    "starac": "starca",
    "udarac": "udarca",
    "trenutak": "trenutka",
    "sastanak": "sastanka",
    "početak": "početka",
    "doručak": "doručka",
    "ručak": "ručka",
    "ulazak": "ulaska",
    "izlazak": "izlaska",
    "dolazak": "dolaska",
    "polazak": "polaska",
    "boraca": "borca",
    "gubitak": "gubitka",
    "dobitak": "dobitka",
    "spavanje": None,  # placeholder, hat keinen Gen-Hinweis nötig
    "lonac": "lonca",
    "lakat": "lakta",
    "nokat": "nokta",
    "palac": "palca",
    "venac": "venca",
    "kupac": "kupca",
    "stranac": "stranca",
    "borac": "borca",
    "vrabac": "vrapca",
    "vepar": "vepra",
    "popa": None,
    "đak": "đaka",  # nicht beweglich
    "izbor": "izbora",  # nicht beweglich
    "trgovac": "trgovca",
    "ljubavnik": "ljubavnika",  # nicht beweglich
    "muzika": None,
    "tovar": "tovara",  # nicht beweglich
    "razlog": "razloga",  # nicht beweglich
    "prozor": "prozora",  # nicht beweglich
}

# Bereinige NOUN_GEN: nur Einträge mit beweglichem Vokal behalten
NOUN_GEN = {k: v for k, v in NOUN_GEN.items() if v is not None and v != k + "a"}


def gen_noun(word):
    """Returns dict with {genus, plural?, gen?}"""
    out = {}
    # Genus
    if word in NEUT_PL_TANTUM:
        out["genus"] = "n"
    elif word in FEM_PL_TANTUM:
        out["genus"] = "f"
    elif word in MASC_ON_A:
        out["genus"] = "m"
    elif word in MASC_ON_O:
        out["genus"] = "m"
        gen_form = MASC_ON_O[word]
        if gen_form:
            out["gen"] = gen_form
    elif word in FEM_ON_CONS:
        out["genus"] = "f"
    elif word.endswith("a"):
        out["genus"] = "f"
    elif word.endswith("o") or word.endswith("e"):
        out["genus"] = "n"
    else:
        # Konsonant: m default, plus -ost/-nost ausnahme
        if word.endswith("ost") or word.endswith("nost"):
            out["genus"] = "f"
        else:
            out["genus"] = "m"

    # Plural unregelmäßig
    if word in NOUN_PLURAL:
        out["plural"] = NOUN_PLURAL[word]

    # Genitiv-Singular bei beweglichem Vokal
    if word in NOUN_GEN and "gen" not in out:
        out["gen"] = NOUN_GEN[word]

    return out


# =====================================================================
# MAIN
# =====================================================================
def main():
    repo = Path(__file__).resolve().parent.parent
    xlsx = repo / "source" / "wortliste.xlsx"
    forms_path = repo / "data" / "forms.json"

    wb = load_workbook(xlsx, read_only=True)
    ws = wb["Wortliste_V2"]
    rows = list(ws.iter_rows(values_only=True))

    existing = {}
    if forms_path.exists():
        with forms_path.open() as f:
            existing = json.load(f)

    adj_generated = 0
    noun_generated = 0
    verb_generated = 0
    for r in rows[1:]:
        if len(r) < 11:
            continue
        nr, de, sl, sc, fil, wa, *_ = r
        if not (wa and de and sl):
            continue
        key = f"{de}|{sl}"
        if key in existing:
            continue
        first = str(sl).split(" / ")[0].strip()
        if not first:
            continue

        if "Adjektiv" in str(wa):
            existing[key] = {"mfn": gen_adj_mfn(first)}
            adj_generated += 1
        elif "Nomen" in str(wa):
            existing[key] = gen_noun(first)
            noun_generated += 1
        elif "Verb" in str(wa):
            # Erkennt: Singleton vs Aspektpaar via "/"
            sl_parts = [p.strip() for p in str(sl).split("/")]
            entry = {}
            if len(sl_parts) >= 2:
                ipf_raw, pf_raw = sl_parts[0], sl_parts[1]
                ipf_stem, _ = strip_se(ipf_raw)
                pf_stem, _ = strip_se(pf_raw)
                ipf_pres = gen_verb_pres1(ipf_stem)
                pf_pres = gen_verb_pres1(pf_stem)
                # Reflexives "se" wieder anhängen, wenn vorhanden
                if ipf_pres and ipf_raw.endswith(" se"):
                    ipf_pres += " se"
                if pf_pres and pf_raw.endswith(" se"):
                    pf_pres += " se"
                if ipf_pres:
                    entry["pres1_ipf"] = ipf_pres
                if pf_pres:
                    entry["pres1_pf"] = pf_pres
            else:
                stem, _ = strip_se(first)
                pres = gen_verb_pres1(stem)
                if pres and first.endswith(" se"):
                    pres += " se"
                if pres:
                    entry["pres1_ipf"] = pres
            if entry:
                existing[key] = entry
                verb_generated += 1

    forms_path.parent.mkdir(parents=True, exist_ok=True)
    with forms_path.open("w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    print(f"Adjektive: {adj_generated} neu generiert.")
    print(f"Nomen:     {noun_generated} neu generiert.")
    print(f"Verben:    {verb_generated} neu generiert.")
    print(f"Total in forms.json: {len(existing)}")


if __name__ == "__main__":
    main()
