"""
Konvertiert source/wortliste.xlsx zu data/words.json
Wird von GitHub Action bei jedem Push auf source/*.xlsx automatisch ausgeführt.

Output Schema:
{
  "version": "<sha256-hash-of-xlsx>",
  "generated": "<ISO timestamp>",
  "themen": ["Arbeit & Organisation", ...],
  "words": [
    {
      "id": <int row-based unique>,
      "n": <Nr from xlsx>,
      "de": "<deutsch>",
      "sl": "<serbisch latinica>",
      "sc": "<serbisch cirilica>",
      "f": "a|b|c",
      "wa": "<Wortart>",
      "fm": "<Form/Hinweis>",
      "tb": "<Themenblock>"
    },
    ...
  ]
}
"""
import hashlib
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


def main(xlsx_path: Path, out_dir: Path):
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    # Hash für Versionierung
    h = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()[:12]

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Wortliste_V2"]
    rows = list(ws.iter_rows(values_only=True))

    words = []
    for idx, r in enumerate(rows[1:]):
        if len(r) < 11:
            continue
        nr, de, sl, sc, fil, wa, form, thema, sf, stufe, tb = r[:11]
        if not de or not sl:
            continue
        words.append({
            "id": idx,
            "n": nr,
            "de": str(de),
            "sl": str(sl),
            "sc": str(sc) if sc else "",
            "f": (str(fil) or "c").lower(),
            "wa": str(wa) if wa else "",
            "fm": str(form) if form else "",
            "tb": str(tb) if tb else "Allgemein",
        })

    themen = sorted({w["tb"] for w in words})

    output = {
        "version": h,
        "generated": datetime.now(timezone.utc).isoformat(),
        "themen": themen,
        "words": words,
    }

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "words.json"
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))

    print(f"Wrote {out_path} ({len(words)} words, version {h})")


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent
    xlsx = root / "source" / "wortliste.xlsx"
    out = root / "data"
    main(xlsx, out)
